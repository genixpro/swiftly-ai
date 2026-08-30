import json
from io import BytesIO
from pathlib import Path

import mongomock
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app import main
from app.main import REPORT_TITLES, report_rows
from app.services.seeding import seed_demo_files
from app.settings import settings


FIXTURES = Path(__file__).parents[1] / "fixtures"


@pytest.fixture
def client(tmp_path, monkeypatch):
    """Exercise both API names against an isolated Mongo-compatible store."""
    monkeypatch.setenv("DATA_DIR", str(tmp_path / "data"))
    # An explicit empty value must override any developer key loaded from the
    # repository-root .env file.
    monkeypatch.setenv("OPENAI_API_KEY", "")
    settings.cache_clear()
    monkeypatch.setattr(main, "MongoClient", mongomock.MongoClient)
    with TestClient(main.app) as test_client:
        yield test_client
    settings.cache_clear()


def test_appraisal_children_and_exports_keep_their_response_shapes(client):
    assert client.get("/health").json() == {"status": "ok"}
    names = {item["name"] for item in client.get("/appraisals/").json()["appraisals"]}
    assert {"Harbour Centre Demo", "Market Hall Demo"} <= names
    seeded = client.get("/appraisals/demo-appraisal").json()["appraisal"]
    assert seeded["units"] and seeded["stabilizedStatement"]["valuation"] > 0
    assert seeded["directComparisonValuation"]["valuation"] > 0
    assert seeded["directComparisonValuation"]["comparativeValue"] == 3_900_000
    assert seeded["validationResult"]["hasBuildingInformation"] is True
    assert seeded["validationResult"]["hasRentRoll"] is True
    assert seeded["validationResult"]["hasFinancialInfo"] is True
    assert seeded["location"] == {"type": "Point", "coordinates": [-79.3777, 43.6426]}
    assert client.get(f"/zones/{seeded['zoning']}").json()["zone"]["zoneName"] == "CR 3.0"
    assert seeded["dataTypeReferences"]["EXPENSE_STATEMENT"][0] == {
        "appraisalId": "demo-appraisal", "fileId": "demo-financial-statement",
        "pageNumbers": [1], "wordIndexes": [],
    }
    seeded_files = client.get("/appraisals/demo-appraisal/files").json()["files"]
    assert {item["_id"] for item in seeded_files} >= {"demo-financial-statement", "demo-lease", "demo-comparable-sale", "demo-scanned-rent-roll"}
    seeded_sale = client.get("/comparable-sales/demo-sale").json()["comparableSale"]
    assert seeded_sale["capitalizationRate"] == 5.25
    assert seeded_sale["occupancyRate"] == 94
    seeded_lease = client.get("/comparable-leases/demo-lease").json()["comparableLease"]
    assert seeded_lease["rentEscalations"][0] == {"startYear": 1, "endYear": 5, "yearlyRent": 160_000}
    assert seeded_lease["taxesMaintenanceInsurance"] == 12.5

    created = client.post("/appraisals/", json={"name": "Contract property", "address": "1 Test Way"})
    assert created.status_code == 200
    appraisal_id = created.json()["_id"]

    updated = client.patch(f"/appraisals/{appraisal_id}", json={"units": [{"tenantName": "Acme", "yearlyRent": 100_000}]})
    assert updated.status_code == 200
    # The default stabilized statement applies the configured 5% vacancy assumption.
    assert updated.json()["appraisal"]["stabilizedStatement"]["netOperatingIncome"] == 95_000
    assert any(item["_id"] == appraisal_id for item in client.get("/appraisals/").json()["appraisals"])

    lease = client.post(f"/appraisals/{appraisal_id}/leases", json={"tenantName": "Acme"}).json()["_id"]
    assert client.get(f"/appraisals/{appraisal_id}/leases/{lease}").json()["lease"]["tenantName"] == "Acme"
    assert client.patch(f"/appraisals/{appraisal_id}/leases/{lease}", json={"tenantName": "Acme 2"}).json() == {"_id": lease}

    statement = client.post(f"/appraisals/{appraisal_id}/financial-statements", json={"year": 2025}).json()["_id"]
    assert client.get(f"/appraisals/{appraisal_id}/financial-statements/{statement}").json()["financialStatement"]["year"] == 2025

    export = client.get(f"/appraisals/{appraisal_id}/reports/rent_roll?format=xlsx")
    assert export.status_code == 200
    assert export.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument")


def test_seeded_file_upgrade_normalizes_legacy_word_tokens(client):
    repositories = client.app.state.repositories
    repositories.files.update_one(
        {"_id": "demo-financial-statement"},
        {"$set": {"words": ["legacy", "tokens"]}},
    )

    seed_demo_files(client.app, FIXTURES)

    words = repositories.files.find_one({"_id": "demo-financial-statement"})["words"]
    assert [word["word"] for word in words] == ["legacy", "tokens"]
    assert all(word["page"] == 1 for word in words)


def test_comparable_zone_tag_and_tenant_contracts(client):
    sale_id = client.post("/comparable-sales", json={"address": "2 Sale Road", "salePrice": 1_500_000}).json()["_id"]
    sales = client.get("/comparable-sales?propertyType=Office").json()
    assert "comparableSales" in sales
    assert client.get(f"/comparable-sales/{sale_id}").json()["comparableSale"]["salePrice"] == 1_500_000
    assert client.patch(f"/comparable-sales/{sale_id}", json={"salePrice": 2_000_000}).json() == {"_id": sale_id}

    lease_id = client.post("/comparable-leases", json={"tenantName": "Northstar", "sizeOfUnit": 1000}).json()["_id"]
    leases = client.get("/comparable-leases?tenantName=Northstar").json()
    assert "comparableLeases" in leases
    assert "Northstar" in client.get("/tenant-names?tenantName=north").json()["names"]
    assert client.get(f"/comparable-leases/{lease_id}").json()["comparableLease"]["sizeOfUnit"] == 1000

    zone_id = client.post("/zones", json={"zoneName": "Downtown"}).json()["_id"]
    assert "zones" in client.get("/zones").json()
    assert client.get(f"/zones/{zone_id}").json()["zone"]["zoneName"] == "Downtown"
    tag_id = client.post("/property-tags", json={"name": "Transit", "propertyType": "Office"}).json()["_id"]
    assert client.get("/property-tags?propertyType=Office").json()["tags"][0]["_id"] == tag_id


def test_image_and_comparable_import_contracts(client):
    image = client.post("/images", files={"file": ("building.png", b"png", "image/png")})
    assert image.status_code == 200
    image_url = image.json()["url"]
    assert image_url.endswith("/images/" + image_url.rsplit("/", 1)[1])
    served = client.get(image_url)
    assert served.status_code == 200 and served.content == b"png"

    imported = client.post(
        "/comparable-sales/import",
        files={"file": ("sales.csv", b"address,sale_price,size_square_footage" + bytes([10]) + b"5 Import Way,$1200000,10000" + bytes([10]), "text/csv")},
    )
    assert imported.status_code == 200
    row = imported.json()["comparableSales"][0]
    assert row["address"] == "5 Import Way"
    assert row["salePrice"] == 1_200_000 and row["sizeSquareFootage"] == 10_000

    extraction_import = client.post(
        "/comparable-sales/import",
        files={"file": ("extraction.json", (FIXTURES / "comparable-sale.json").read_bytes(), "application/json")},
    )
    extracted = extraction_import.json()["comparableSales"][0]
    assert extracted["address"] == "25 King Street, Toronto, Ontario"
    assert extracted["salePrice"] == 12_500_000 and extracted["sizeSquareFootage"] == 50_000


def test_canonical_mutation_routes(client):
    appraisal_id = client.post("/appraisals/", json={"name": "Route coverage", "address": "3 Test Way"}).json()["_id"]
    assert client.get("/appraisals").status_code == 200
    client.patch(f"/appraisals/{appraisal_id}", json={"units": [{"tenantName": "Acme", "yearlyRent": 100_000}]})
    assert client.post(f"/appraisals/{appraisal_id}/comparable-leases/from-tenants").json()["created"]

    lease = client.post(f"/appraisals/{appraisal_id}/leases", json={"tenantName": "Lease row"}).json()["_id"]
    assert len(client.get(f"/appraisals/{appraisal_id}/leases").json()["leases"]) == 1
    statement = client.post(f"/appraisals/{appraisal_id}/financial-statements", json={"year": 2025}).json()["_id"]
    assert len(client.get(f"/appraisals/{appraisal_id}/financial-statements").json()["financial_statements"]) == 1
    assert client.patch(f"/appraisals/{appraisal_id}/financial-statements/{statement}", json={"year": 2026}).json() == {"_id": statement}
    assert client.patch(f"/appraisals/{appraisal_id}/leases/{lease}", json={"tenantName": "Changed"}).json() == {"_id": lease}

    comparable = client.post("/comparable-leases", json={"tenantName": "Coverage tenant", "sizeOfUnit": 500}).json()["_id"]
    assert client.patch(f"/comparable-leases/{comparable}", json={"sizeOfUnit": 600}).json() == {"_id": comparable}
    assert client.delete(f"/comparable-leases/{comparable}").json() == {}
    sale = client.post("/comparable-sales", json={"address": "3 Sale Road"}).json()["_id"]
    portfolio = client.post("/comparable-sale-portfolios", json={"portfolio": {"address": "Portfolio"}, "subComps": [{"address": "Subcomp"}]}).json()
    assert portfolio["subCompIds"] and client.delete(f"/comparable-sales/{sale}").json() == {}
    singular_sale = client.post("/comparable-sales", json={"address": "Singular delete Road"}).json()["_id"]
    assert client.delete(f"/comparable-sales/{singular_sale}").json() == {}

    zone = client.post("/zones", json={"zoneName": "Coverage zone"}).json()["_id"]
    assert client.patch(f"/zones/{zone}", json={"description": "Updated"}).json() == {"_id": zone}
    assert client.delete(f"/zones/{zone}").json() == {}
    tag = client.post("/property-tags", json={"name": "Coverage tag"}).json()["_id"]
    assert client.delete(f"/property-tags/{tag}").json() == {}
    assert client.delete(f"/appraisals/{appraisal_id}").json() == {}


@pytest.mark.parametrize(
    ("method", "path"),
    [
        ("get", "/appraisal/"),
        ("post", "/appraisal/"),
        ("get", "/appraisal/demo-appraisal"),
        ("get", "/comparable_sales"),
        ("get", "/comparable_leases"),
        ("get", "/property_tags"),
        ("get", "/tenant_names"),
        ("get", "/zone/demo-zone"),
        ("post", "/comparable_sale_upload/"),
    ],
)
def test_removed_legacy_routes_return_not_found(client, method, path):
    response = getattr(client, method)(path)
    assert response.status_code == 404


def test_file_upload_extract_and_manual_extraction_correction_contracts(client):
    missing_appraisal = client.post(
        "/appraisals/missing/files",
        files={"file": ("financial-statement.pdf", (FIXTURES / "financial-statement.pdf").read_bytes(), "application/pdf")},
    )
    assert missing_appraisal.status_code == 404

    modern_upload = client.post(
        "/appraisals/demo-appraisal/files",
        files={"file": ("scan.png", (FIXTURES / "image-placeholder.png").read_bytes(), "image/png")},
    )
    assert modern_upload.status_code == 200
    modern_file = modern_upload.json()["file"]
    assert modern_file["reviewStatus"] == "queued"
    modern_job = client.get(f"/extractions/{modern_file['extractionJobId']}").json()["extraction"]
    assert modern_job["status"] == "failed"
    assert client.get(f"/appraisals/demo-appraisal/files/{modern_file['_id']}/rendered-pages/1").headers["content-type"] == "image/png"
    extraction = client.post(f"/appraisals/demo-appraisal/files/{modern_file['_id']}/extract")
    assert extraction.status_code == 200 and extraction.json()["status"] == "queued"

    upload = client.post(
        "/appraisals/demo-appraisal/files",
        files={"file": ("financial-statement.pdf", (FIXTURES / "financial-statement.pdf").read_bytes(), "application/pdf")},
    )
    assert upload.status_code == 200
    file = upload.json()["file"]
    assert file["fileName"] == "financial-statement.pdf"
    assert file["extractionJobId"]

    job = client.get(f"/extractions/{file['extractionJobId']}").json()["extraction"]
    assert job["status"] == "failed"  # no runtime key: failure remains reviewable/reprocessable
    assert "OPENAI_API_KEY" in job["error"]
    reviewed = client.get(f"/appraisals/demo-appraisal/files/{file['_id']}").json()["file"]
    assert reviewed["pages"] > 0
    assert "path" not in reviewed and "images" not in reviewed
    assert reviewed["reviewStatus"] == "extraction_failed"
    assert "OPENAI_API_KEY" in reviewed["extractionError"]
    listed = client.get("/appraisals/demo-appraisal/files").json()["files"]
    assert next(item for item in listed if item["_id"] == file["_id"])["extractionJobId"] == file["extractionJobId"]
    rendered = client.get(f"/appraisals/demo-appraisal/files/{file['_id']}/rendered-pages/1")
    assert rendered.status_code == 200 and rendered.headers["content-type"] == "image/png"
    contents = client.get(f"/appraisals/demo-appraisal/files/{file['_id']}/content")
    assert contents.status_code == 200 and contents.content[:4] == b"%PDF"

    correction = {
        "extraction": {
            "document_type": "financial_statement", "confidence": 0.8,
            "fields": [{"name": "period", "value": "2025", "citations": [{"page": 1, "text": "Statement"}]}],
            "tenant_lease_rows": [], "income_expense_rows": [], "comparable_sales": [],
            "citations": [{"page": 1, "text": "Statement"}],
        }
    }
    patched = client.patch(f"/appraisals/demo-appraisal/files/{file['_id']}/extraction", json=correction)
    assert patched.status_code == 200
    patched_file = patched.json()["file"]
    assert patched_file["reviewStatus"] == "corrected"
    assert patched_file["extraction"] == correction["extraction"]
    assert "annotations" not in patched_file
    assert "path" not in patched_file and "images" not in patched_file and "hash" not in patched_file
    unsafe_patch = client.patch(
        f"/appraisals/demo-appraisal/files/{file['_id']}",
        json={"path": "/etc/hosts"},
    )
    assert unsafe_patch.status_code == 422
    classified = client.patch(
        f"/appraisals/demo-appraisal/files/{file['_id']}",
        json={"fileType": "financials"},
    )
    assert classified.status_code == 200
    assert classified.json()["file"]["fileType"] == "financials"
    retry = client.post(f"/appraisals/demo-appraisal/files/{file['_id']}/extract")
    assert retry.status_code == 200
    retried_file = client.get(f"/appraisals/demo-appraisal/files/{file['_id']}").json()["file"]
    assert retried_file["extractionJobId"] == retry.json()["jobId"]
    assert retried_file["reviewStatus"] == "extraction_failed"
    source_paths = list((settings().data_dir / "uploads").glob(f"{file['_id']}.*"))
    rendered_paths = list((settings().data_dir / "rendered" / file["_id"]).glob("page-*.png"))
    assert len(source_paths) == 1 and source_paths[0].is_file()
    assert rendered_paths and rendered_paths[0].is_file()
    assert client.delete(f"/appraisals/demo-appraisal/files/{file['_id']}").json() == {}
    assert not source_paths[0].exists() and not rendered_paths[0].exists()
    assert client.get(f"/extractions/{retry.json()['jobId']}").status_code == 404


def test_report_mappers_and_downloads_match_versioned_golden_output(client):
    golden = json.loads((FIXTURES / "export-golden.json").read_text())
    appraisal = golden["appraisal"]
    assert {
        report: [list(row) for row in report_rows(appraisal, report)]
        for report in golden["expectedRows"]
    } == golden["expectedRows"]

    appraisal_id = client.post("/appraisals/", json=appraisal).json()["_id"]
    for report in REPORT_TITLES:
        for format in ("docx", "xlsx"):
            response = client.get(f"/appraisals/{appraisal_id}/reports/{report}?format={format}")
            assert response.status_code == 200
            assert response.content[:2] == b"PK"  # .docx/.xlsx zip containers
    detailed = client.get(f"/appraisals/{appraisal_id}/reports/comparable_sales?format=detailed-docx")
    assert detailed.status_code == 200 and detailed.content[:2] == b"PK"


def test_reports_resolve_comparable_ids_saved_by_the_react_screens(client):
    sale_id = client.post("/comparable-sales", json={"address": "Linked Sale", "salePrice": 2_400_000}).json()["_id"]
    lease_id = client.post("/comparable-leases", json={"tenantName": "Linked Lease", "sizeOfUnit": 3600}).json()["_id"]
    appraisal_id = client.post("/appraisals/", json={
        "name": "Linked comparables", "address": "4 Export Way",
        "comparableSalesCapRate": [sale_id], "comparableSalesDCA": [sale_id], "comparableLeases": [lease_id],
    }).json()["_id"]

    sales = load_workbook(BytesIO(client.get(f"/appraisals/{appraisal_id}/reports/comparable_sales?format=xlsx").content)).active
    leases = load_workbook(BytesIO(client.get(f"/appraisals/{appraisal_id}/reports/comparable_leases?format=xlsx").content)).active
    assert list(sales.values)[1] == ("Linked Sale", 2_400_000)
    assert list(leases.values)[1] == ("Linked Lease", 3600)
