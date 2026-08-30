import json
from io import BytesIO
from pathlib import Path

import mongomock
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app import main
from app.main import REPORT_TITLES, report_rows
from app.settings import settings


FIXTURES = Path(__file__).parents[1] / "fixtures"


@pytest.fixture
def client(tmp_path, monkeypatch):
    """Exercise both API names against an isolated Mongo-compatible store."""
    monkeypatch.setenv("DATA_DIR", str(tmp_path / "data"))
    # An explicit empty value must override any developer key loaded from the
    # repository-root .env file.
    monkeypatch.setenv("OPENAI_API_KEY", "")
    settings.cache_clear()
    monkeypatch.setattr(main, "MongoClient", mongomock.MongoClient)
    with TestClient(main.app) as test_client:
        yield test_client
    settings.cache_clear()


def test_legacy_appraisal_children_and_exports_keep_their_response_shapes(client):
    assert client.get("/health").json() == {"status": "ok"}
    names = {item["name"] for item in client.get("/appraisal/").json()["appraisals"]}
    assert {"Harbour Centre Demo", "Market Hall Demo"} <= names
    seeded = client.get("/appraisal/demo-appraisal").json()["appraisal"]
    assert seeded["units"] and seeded["stabilizedStatement"]["valuation"] > 0
    assert seeded["directComparisonValuation"]["valuation"] > 0
    assert seeded["directComparisonValuation"]["comparativeValue"] == 3_900_000
    assert seeded["validationResult"]["hasBuildingInformation"] is True
    assert seeded["validationResult"]["hasRentRoll"] is True
    assert seeded["validationResult"]["hasFinancialInfo"] is True
    assert seeded["location"] == {"type": "Point", "coordinates": [-79.3777, 43.6426]}
    assert client.get(f"/zone/{seeded['zoning']}").json()["zone"]["zoneName"] == "CR 3.0"
    assert seeded["dataTypeReferences"]["EXPENSE_STATEMENT"][0] == {
        "appraisalId": "demo-appraisal", "fileId": "demo-financial-statement",
        "pageNumbers": [1], "wordIndexes": [],
    }
    seeded_files = client.get("/appraisal/demo-appraisal/files").json()["files"]
    assert {item["_id"] for item in seeded_files} >= {"demo-financial-statement", "demo-lease", "demo-comparable-sale", "demo-scanned-rent-roll"}
    seeded_sale = client.get("/comparable_sales/demo-sale").json()["comparableSale"]
    assert seeded_sale["capitalizationRate"] == 5.25
    assert seeded_sale["occupancyRate"] == 94
    seeded_lease = client.get("/comparable_leases/demo-lease").json()["comparableLease"]
    assert seeded_lease["rentEscalations"][0] == {"startYear": 1, "endYear": 5, "yearlyRent": 160_000}
    assert seeded_lease["taxesMaintenanceInsurance"] == 12.5

    created = client.post("/appraisal/", json={"name": "Contract property", "address": "1 Test Way"})
    assert created.status_code == 200
    appraisal_id = created.json()["_id"]

    updated = client.post(f"/appraisal/{appraisal_id}", json={"units": [{"tenantName": "Acme", "yearlyRent": 100_000}]})
    assert updated.status_code == 200
    # The default stabilized statement applies the configured 5% vacancy assumption.
    assert updated.json()["appraisal"]["stabilizedStatement"]["netOperatingIncome"] == 95_000
    assert any(item["_id"] == appraisal_id for item in client.get("/appraisal/").json()["appraisals"])

    lease = client.post(f"/appraisal/{appraisal_id}/leases", json={"tenantName": "Acme"}).json()["_id"]
    assert client.get(f"/appraisal/{appraisal_id}/leases/{lease}").json()["lease"]["tenantName"] == "Acme"
    assert client.post(f"/appraisal/{appraisal_id}/leases/{lease}", json={"tenantName": "Acme 2"}).json() == {"_id": lease}

    statement = client.post(f"/appraisal/{appraisal_id}/financial_statements", json={"year": 2025}).json()["_id"]
    assert client.get(f"/appraisal/{appraisal_id}/financial_statements/{statement}").json()["financialStatement"]["year"] == 2025

    export = client.get(f"/appraisal/{appraisal_id}/rent_roll/excel")
    assert export.status_code == 200
    assert export.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument")


def test_legacy_comparable_zone_tag_and_tenant_contracts(client):
    sale_id = client.post("/comparable_sales", json={"address": "2 Sale Road", "salePrice": 1_500_000}).json()["_id"]
    sales = client.get("/comparable_sales?propertyType=Office").json()
    assert "comparableSales" in sales
    assert client.get(f"/comparable_sales/{sale_id}").json()["comparableSale"]["salePrice"] == 1_500_000
    assert client.post(f"/comparable_sales/{sale_id}", json={"salePrice": 2_000_000}).json() == {"_id": sale_id}

    lease_id = client.post("/comparable_leases", json={"tenantName": "Northstar", "sizeOfUnit": 1000}).json()["_id"]
    leases = client.get("/comparable_leases?tenantName=Northstar").json()
    assert "comparableLeases" in leases
    assert "Northstar" in client.get("/tenant_names?tenantName=north").json()["names"]
    assert client.get(f"/comparable_leases/{lease_id}").json()["comparableLease"]["sizeOfUnit"] == 1000

    zone_id = client.post("/zones", json={"zoneName": "Downtown"}).json()["_id"]
    assert "zones" in client.get("/zones").json()
    assert client.get(f"/zone/{zone_id}").json()["zone"]["zoneName"] == "Downtown"
    tag_id = client.post("/property_tags", json={"name": "Transit", "propertyType": "Office"}).json()["_id"]
    assert client.get("/property_tags?propertyType=Office").json()["tags"][0]["_id"] == tag_id


def test_legacy_image_and_comparable_import_contracts(client):
    image = client.post("/images", files={"file": ("building.png", b"png", "image/png")})
    assert image.status_code == 200
    image_url = image.json()["url"]
    assert image_url.endswith("/images/" + image_url.rsplit("/", 1)[1])
    served = client.get(image_url)
    assert served.status_code == 200 and served.content == b"png"

    imported = client.post(
        "/comparable_sale_upload/",
        files={"file": ("sales.csv", b"address,sale_price,size_square_footage" + bytes([10]) + b"5 Import Way,$1200000,10000" + bytes([10]), "text/csv")},
    )
    assert imported.status_code == 200
    row = imported.json()["comparableSales"][0]
    assert row["address"] == "5 Import Way"
    assert row["salePrice"] == 1_200_000 and row["sizeSquareFootage"] == 10_000

    extraction_import = client.post(
        "/comparable_sale_upload/",
        files={"file": ("extraction.json", (FIXTURES / "comparable-sale.json").read_bytes(), "application/json")},
    )
    extracted = extraction_import.json()["comparableSales"][0]
    assert extracted["address"] == "25 King Street, Toronto, Ontario"
    assert extracted["salePrice"] == 12_500_000 and extracted["sizeSquareFootage"] == 50_000


def test_remaining_legacy_mutation_adapters(client):
    appraisal_id = client.post("/appraisal/", json={"name": "Route coverage", "address": "3 Test Way"}).json()["_id"]
    assert client.get("/appraisals").status_code == 200
    client.post(f"/appraisal/{appraisal_id}", json={"units": [{"tenantName": "Acme", "yearlyRent": 100_000}]})
    assert client.post(f"/appraisal/{appraisal_id}/convert_tenants").json()["created"]

    lease = client.post(f"/appraisal/{appraisal_id}/leases", json={"tenantName": "Lease row"}).json()["_id"]
    assert len(client.get(f"/appraisal/{appraisal_id}/leases").json()["leases"]) == 1
    statement = client.post(f"/appraisal/{appraisal_id}/financial_statements", json={"year": 2025}).json()["_id"]
    assert len(client.get(f"/appraisal/{appraisal_id}/financial_statements").json()["financial_statements"]) == 1
    assert client.post(f"/appraisal/{appraisal_id}/financial_statements/{statement}", json={"year": 2026}).json() == {"_id": statement}
    assert client.post(f"/appraisal/{appraisal_id}/leases/{lease}", json={"tenantName": "Changed"}).json() == {"_id": lease}

    comparable = client.post("/comparable_leases", json={"tenantName": "Coverage tenant", "sizeOfUnit": 500}).json()["_id"]
    assert client.post(f"/comparable_leases/{comparable}", json={"sizeOfUnit": 600}).json() == {"_id": comparable}
    assert client.delete(f"/comparable_leases/{comparable}").json() == {}
    sale = client.post("/comparable_sales", json={"address": "3 Sale Road"}).json()["_id"]
    portfolio = client.post("/comparable_sales_portfolio/", json={"portfolio": {"address": "Portfolio"}, "subComps": [{"address": "Subcomp"}]}).json()
    assert portfolio["subCompIds"] and client.delete(f"/comparable_sales/{sale}").json() == {}
    singular_sale = client.post("/comparable_sales", json={"address": "Singular delete Road"}).json()["_id"]
    assert client.delete(f"/comparable_sale/{singular_sale}").json() == {}

    zone = client.post("/zones", json={"zoneName": "Coverage zone"}).json()["_id"]
    assert client.post(f"/zone/{zone}", json={"description": "Updated"}).json() == {"_id": zone}
    assert client.delete(f"/zone/{zone}").json() == {}
    tag = client.post("/property_tags", json={"name": "Coverage tag"}).json()["_id"]
    assert client.delete(f"/property_tags/{tag}").json() == {}
    assert client.delete(f"/appraisal/{appraisal_id}").json() == {}


def test_file_upload_reprocess_and_manual_extraction_correction_contracts(client):
    missing_appraisal = client.post(
        "/appraisals/missing/files",
        files={"file": ("financial-statement.pdf", (FIXTURES / "financial-statement.pdf").read_bytes(), "application/pdf")},
    )
    assert missing_appraisal.status_code == 404

    modern_upload = client.post(
        "/appraisals/demo-appraisal/files",
        files={"file": ("scan.png", (FIXTURES / "image-placeholder.png").read_bytes(), "image/png")},
    )
    assert modern_upload.status_code == 200
    modern_file = modern_upload.json()["file"]
    assert modern_file["reviewStatus"] == "queued"
    modern_job = client.get(f"/extractions/{modern_file['extractionJobId']}").json()["extraction"]
    assert modern_job["status"] == "failed"
    assert client.get(f"/appraisal/demo-appraisal/files/{modern_file['_id']}/rendered/1").headers["content-type"] == "image/png"
    modern_reprocess = client.post(f"/appraisals/demo-appraisal/files/{modern_file['_id']}/extract")
    assert modern_reprocess.status_code == 200 and modern_reprocess.json()["status"] == "queued"

    upload = client.post(
        "/appraisal/demo-appraisal/files",
        files={"file": ("financial-statement.pdf", (FIXTURES / "financial-statement.pdf").read_bytes(), "application/pdf")},
    )
    assert upload.status_code == 200
    file = upload.json()["file"]
    assert file["fileName"] == "financial-statement.pdf"
    assert file["extractionJobId"]

    job = client.get(f"/extractions/{file['extractionJobId']}").json()["extraction"]
    assert job["status"] == "failed"  # no runtime key: failure remains reviewable/reprocessable
    assert "OPENAI_API_KEY" in job["error"]
    reviewed = client.get(f"/appraisal/demo-appraisal/files/{file['_id']}").json()["file"]
    assert reviewed["pages"] > 0 and reviewed["images"]
    assert reviewed["reviewStatus"] == "extraction_failed"
    assert "OPENAI_API_KEY" in reviewed["extractionError"]
    listed = client.get("/appraisal/demo-appraisal/files").json()["files"]
    assert next(item for item in listed if item["_id"] == file["_id"])["extractionJobId"] == file["extractionJobId"]
    rendered = client.get(f"/appraisal/demo-appraisal/files/{file['_id']}/rendered/1")
    assert rendered.status_code == 200 and rendered.headers["content-type"] == "image/png"
    contents = client.get(f"/appraisal/demo-appraisal/files/{file['_id']}/contents")
    assert contents.status_code == 200 and contents.content[:4] == b"%PDF"

    correction = {
        "extraction": {
            "document_type": "financial_statement", "confidence": 0.8,
            "fields": [{"name": "period", "value": "2025", "citations": [{"page": 1, "text": "Statement"}]}],
            "tenant_lease_rows": [], "income_expense_rows": [], "comparable_sales": [],
            "citations": [{"page": 1, "text": "Statement"}],
        }
    }
    patched = client.patch(f"/appraisals/demo-appraisal/files/{file['_id']}/extraction", json=correction)
    assert patched.status_code == 200
    patched_file = patched.json()["file"]
    assert patched_file["reviewStatus"] == "corrected"
    assert patched_file["extraction"] == correction["extraction"]
    assert "annotations" not in patched_file
    reprocess = client.post(f"/appraisal/demo-appraisal/files/{file['_id']}/reprocess")
    assert reprocess.status_code == 200
    reprocessed_file = client.get(f"/appraisal/demo-appraisal/files/{file['_id']}").json()["file"]
    assert reprocessed_file["extractionJobId"] == reprocess.json()["jobId"]
    assert reprocessed_file["reviewStatus"] == "extraction_failed"
    source_path = Path(reprocessed_file["path"])
    rendered_path = Path(reprocessed_file["images"][0])
    assert source_path.is_file() and rendered_path.is_file()
    assert client.delete(f"/appraisal/demo-appraisal/files/{file['_id']}").json() == {}
    assert not source_path.exists() and not rendered_path.exists()
    assert client.get(f"/extractions/{reprocess.json()['jobId']}").status_code == 404


def test_report_mappers_and_downloads_match_versioned_golden_output(client):
    golden = json.loads((FIXTURES / "export-golden.json").read_text())
    appraisal = golden["appraisal"]
    assert {
        report: [list(row) for row in report_rows(appraisal, report)]
        for report in golden["expectedRows"]
    } == golden["expectedRows"]

    appraisal_id = client.post("/appraisal/", json=appraisal).json()["_id"]
    for report in REPORT_TITLES:
        for format in ("word", "excel"):
            response = client.get(f"/appraisal/{appraisal_id}/{report}/{format}")
            assert response.status_code == 200
            assert response.content[:2] == b"PK"  # .docx/.xlsx zip containers
    detailed = client.get(f"/appraisal/{appraisal_id}/comparable_sales/detailed_word")
    assert detailed.status_code == 200 and detailed.content[:2] == b"PK"


def test_reports_resolve_comparable_ids_saved_by_the_react_screens(client):
    sale_id = client.post("/comparable_sales", json={"address": "Linked Sale", "salePrice": 2_400_000}).json()["_id"]
    lease_id = client.post("/comparable_leases", json={"tenantName": "Linked Lease", "sizeOfUnit": 3600}).json()["_id"]
    appraisal_id = client.post("/appraisal/", json={
        "name": "Linked comparables", "address": "4 Export Way",
        "comparableSalesCapRate": [sale_id], "comparableSalesDCA": [sale_id], "comparableLeases": [lease_id],
    }).json()["_id"]

    sales = load_workbook(BytesIO(client.get(f"/appraisal/{appraisal_id}/comparable_sales/excel").content)).active
    leases = load_workbook(BytesIO(client.get(f"/appraisal/{appraisal_id}/comparable_leases/excel").content)).active
    assert list(sales.values)[1] == ("Linked Sale", 2_400_000)
    assert list(leases.values)[1] == ("Linked Lease", 3600)
